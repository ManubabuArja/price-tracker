import openpyxl
from openpyxl import Workbook
import os

FILE_PATH = "data/warehouse.xlsx"

# Ensure Excel file exists
def init_excel():
    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Product Name", "URL", "Current Price", "Target Price"])
        wb.save(FILE_PATH)

# Read products from Excel
def read_products():
    init_excel()
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb["Products"]
    products = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            products.append({
                "name": row[0],
                "url": row[1],
                "current_price": row[2],
                "target_price": row[3]
            })
    wb.close()
    return products

# Update product price in Excel
def update_price(product_name, new_price):
    init_excel()
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb["Products"]

    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_name:
            row[2].value = new_price  # Update Current Price
            break

    wb.save(FILE_PATH)
    wb.close()
